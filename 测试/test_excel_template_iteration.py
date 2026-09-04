# -*- coding: utf-8 -*-
"""Excel 模板上传、连续迭代和 Word 工程动向组合的回归测试。"""

import copy
import io
import json
import os
import sys
import tempfile
import unittest
from unittest import mock

import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE)

import app as app_module


class ExcelTemplateIterationTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory(prefix="excel_template_test_")
        self.root = self.temp_dir.name
        self.upload_dir = os.path.join(self.root, "uploaded_templates")
        self.output_dir = os.path.join(self.root, "outputs")
        os.makedirs(self.upload_dir)
        os.makedirs(self.output_dir)

        self.template_path = os.path.join(self.root, "用户最新版模板.xlsx")
        self._make_template(self.template_path)
        with open(os.path.join(BASE, "config.json"), "r", encoding="utf-8") as handle:
            self.config = copy.deepcopy(json.load(handle))
        self.config["word"]["enabled"] = False
        self.config["excel"]["template"] = self.template_path

        self.patchers = [
            mock.patch.object(app_module, "EXCEL_TEMPLATE_DIR", self.upload_dir),
            mock.patch.object(
                app_module, "EXCEL_STATE_PATH",
                os.path.join(self.root, "excel_template_state.json")),
            mock.patch.object(app_module, "load_config", return_value=self.config),
        ]
        for patcher in self.patchers:
            patcher.start()
            self.addCleanup(patcher.stop)

        flask_app = app_module.create_app()
        flask_app.config["TESTING"] = True
        self.client = flask_app.test_client()

    def tearDown(self):
        self.temp_dir.cleanup()

    @staticmethod
    def _make_template(path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "万羿辉"
        sheet.cell(1, 1, "测试迭代模板")
        sheet.cell(2, 1, "序号")
        sheet.cell(2, 2, "施工\n类型")
        sheet.cell(2, 3, "线路名称")
        sheet.cell(2, 4, "设备主人")
        sheet.cell(2, 5, "危险源排摸情况")
        sheet.cell(3, 5, "危险源描述")
        # 即使迭代模板中出现同名列，三个现场选项也不应写入 Excel。
        sheet.cell(3, 6, "是否安全告知")
        sheet.cell(4, 1, 1)
        sheet.cell(4, 3, "模板原记录")
        sheet.cell(4, 4, "万羿辉")
        workbook.save(path)
        workbook.close()

    def _upload_template(self):
        with open(self.template_path, "rb") as handle:
            response = self.client.post(
                "/api/upload_excel_template",
                data={
                    "file": (io.BytesIO(handle.read()), "用户最新版模板.xlsx"),
                    "continue_previous": "true",
                },
                content_type="multipart/form-data",
            )
        self.assertEqual(response.status_code, 200, response.get_data(as_text=True))
        return response.get_json()

    def _generate(self, line_name):
        response = self.client.post("/api/generate", json={
            "fields": {
                "line_name": line_name,
                "equipment_owner": "万羿辉",
                "construction_type": "工地",
                "danger_points": "测试危险点",
                "safety_notice": "是",
                "site_plan_provided": "否",
                "cooperation_level": "一般",
                "form_date": "2026-09-05",
            },
            "continue_excel": True,
            "output_dir": self.output_dir,
        })
        self.assertEqual(response.status_code, 200, response.get_data(as_text=True))
        data = response.get_json()
        output = next(item for item in data["outputs"] if item["type"] == "excel")
        return os.path.join(data["output_folder"], output["name"]), data

    def test_uploaded_template_is_used_and_previous_output_is_reused(self):
        upload = self._upload_template()
        self.assertEqual(upload["excel_template"]["base_source"], "uploaded")
        self.assertEqual(upload["excel_template"]["last_generated_name"], "")

        first_path, first_data = self._generate("第一次追加")
        first_book = openpyxl.load_workbook(first_path, data_only=False)
        first_sheet = first_book["万羿辉"]
        self.assertEqual(first_sheet.cell(5, 2).value, "工地")
        self.assertEqual(first_sheet.cell(5, 3).value, "第一次追加")
        self.assertIn(first_sheet.cell(5, 6).value, (None, ""))
        first_book.close()
        self.assertEqual(
            first_data["excel_template"]["last_generated_name"],
            os.path.basename(first_path))

        # 模拟关闭并重新打开程序，确认迭代状态会从磁盘恢复。
        restarted_app = app_module.create_app()
        restarted_app.config["TESTING"] = True
        self.client = restarted_app.test_client()
        second_path, second_data = self._generate("第二次追加")
        second_book = openpyxl.load_workbook(second_path, data_only=False)
        second_sheet = second_book["万羿辉"]
        self.assertEqual(second_sheet.cell(5, 3).value, "第一次追加")
        self.assertEqual(second_sheet.cell(6, 3).value, "第二次追加")
        self.assertIn(second_sheet.cell(6, 6).value, (None, ""))
        second_book.close()
        self.assertEqual(
            second_data["excel_template"]["last_generated_name"],
            os.path.basename(second_path))

        reset = self.client.post("/api/reset_excel_template").get_json()
        self.assertEqual(reset["excel_template"]["base_source"], "default")
        self.assertEqual(reset["excel_template"]["last_generated_name"], "")

    def test_three_choices_are_composed_only_into_word_trend(self):
        composed = app_module._compose_word_engineering_trend({
            "engineering_trend": "现场正在进行桩基施工。",
            "safety_notice": "留置送达",
            "site_plan_provided": "否",
            "cooperation_level": "一般",
        })
        self.assertEqual(
            composed,
            "现场正在进行桩基施工。\n"
            "是否安全告知：留置送达；施工方是否提供平面图：否；"
            "预估施工方配合程度：一般。")


if __name__ == "__main__":
    unittest.main()
