# -*- coding: utf-8 -*-
"""
后端服务：提供 OCR 识别、附件上传、模板填充、文件下载等接口。

本文件既可以被 main.py 引入（桌面窗口模式），也可以单独运行进行调试：
    python app.py --web     （用浏览器打开界面，便于开发调试）
"""

import datetime
import json
import os
import platform
import re
import sys
import threading
import time
import traceback
import uuid

from flask import (Flask, has_request_context, jsonify, request, send_file,
                   send_from_directory)

from ocr_engine import ocr_engine


# ============================================================
# 路径工具（兼容打包成 exe 后的路径）
# ============================================================
def is_frozen():
    return getattr(sys, "frozen", False)


def base_dir():
    """exe 同目录（存放 config.json、templates、uploads、output）。"""
    if is_frozen():
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def resource_dir():
    """打包后的内置资源目录（static 前端文件等）。"""
    if is_frozen():
        return getattr(sys, "_MEIPASS", base_dir())
    return base_dir()


BASE = base_dir()
STATIC_DIR = os.path.join(resource_dir(), "static")
CONFIG_PATH = os.path.join(BASE, "config.json")
UPLOAD_DIR = os.path.join(BASE, "uploads")
ATTACH_DIR = os.path.join(BASE, "uploads", "attachments")
EXCEL_TEMPLATE_DIR = os.path.join(BASE, "uploads", "excel_templates")
OUTPUT_DIR = os.path.join(BASE, "output")
TEMPLATE_DIR = os.path.join(BASE, "templates")
EXCEL_STATE_PATH = os.path.join(BASE, "excel_template_state.json")

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff", ".webp"}
OUTPUT_FILES = {}
LAST_OUTPUT_DIR = OUTPUT_DIR
EXCEL_STATE_LOCK = threading.RLock()

for _d in (UPLOAD_DIR, ATTACH_DIR, EXCEL_TEMPLATE_DIR, OUTPUT_DIR, TEMPLATE_DIR):
    os.makedirs(_d, exist_ok=True)


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def log_error(context, exc):
    """把接口异常写成带运行环境和上下文的诊断记录。"""
    try:
        with open(os.path.join(BASE, "崩溃日志.log"), "a", encoding="utf-8") as f:
            f.write("=" * 70 + "\n")
            f.write("发生时间：%s\n" % time.strftime("%Y-%m-%d %H:%M:%S"))
            f.write("故障位置：%s\n" % context)
            f.write("程序目录：%s\n" % BASE)
            f.write("运行模式：%s\n" % ("打包版 exe" if is_frozen() else "源码运行"))
            f.write("Python：%s\n" % sys.version.replace("\n", " "))
            f.write("系统：%s\n" % platform.platform())
            if has_request_context():
                f.write("当前接口：%s %s\n" % (request.method, request.path))
            else:
                f.write("当前接口：非 HTTP 接口调用\n")
            f.write("异常类型：%s\n" % type(exc).__name__)
            f.write("异常信息：%s\n\n" % exc)
            f.write("详细堆栈：\n")
            f.write("".join(traceback.format_exception(
                type(exc), exc, exc.__traceback__)))
            f.write("\n排查提示：请把本日志文件完整发回，并同时说明操作步骤。\n")
    except Exception:
        pass


def choose_output_directory():
    """打开 Windows 原生目录选择框；取消时返回空字符串。"""
    if os.name != "nt":
        return ""
    import ctypes
    from ctypes import wintypes

    class BrowseInfo(ctypes.Structure):
        _fields_ = [
            ("hwndOwner", ctypes.c_void_p),
            ("pidlRoot", ctypes.c_void_p),
            ("pszDisplayName", ctypes.c_wchar_p),
            ("lpszTitle", ctypes.c_wchar_p),
            ("ulFlags", ctypes.c_uint),
            ("lpfn", ctypes.c_void_p),
            ("lParam", ctypes.c_void_p),
            ("iImage", ctypes.c_int),
        ]

    shell32 = ctypes.windll.shell32
    ole32 = ctypes.windll.ole32
    shell32.SHBrowseForFolderW.argtypes = [ctypes.POINTER(BrowseInfo)]
    shell32.SHGetPathFromIDListW.argtypes = [ctypes.c_void_p, ctypes.c_wchar_p]
    shell32.SHBrowseForFolderW.restype = ctypes.c_void_p
    shell32.SHGetPathFromIDListW.restype = wintypes.BOOL
    ole32.CoTaskMemFree.argtypes = [ctypes.c_void_p]
    display_name = ctypes.create_unicode_buffer(260)
    info = BrowseInfo(
        0, 0, display_name, "请选择生成文件的输出目录",
        0x0001 | 0x0040, 0, 0, 0,
    )
    pidl = shell32.SHBrowseForFolderW(ctypes.byref(info))
    if not pidl:
        return ""
    try:
        path_buffer = ctypes.create_unicode_buffer(32768)
        if shell32.SHGetPathFromIDListW(pidl, path_buffer):
            return path_buffer.value
        return ""
    finally:
        ole32.CoTaskMemFree(pidl)


def _safe_folder_component(value, fallback):
    value = str(value or "").strip()
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "-", value)
    value = value.strip(" .")
    return value[:80] or fallback


def _empty_excel_state():
    return {
        "uploaded_template": "",
        "uploaded_original": "",
        "last_generated_excel": "",
        "continue_previous": True,
    }


def _load_excel_state():
    """读取 Excel 模板迭代状态；文件损坏时安全回退到内置模板。"""
    state = _empty_excel_state()
    try:
        with open(EXCEL_STATE_PATH, "r", encoding="utf-8") as f:
            saved = json.load(f)
        if not isinstance(saved, dict):
            return state

        stored_name = str(saved.get("uploaded_template") or "").strip()
        if (stored_name == os.path.basename(stored_name) and
                os.path.splitext(stored_name)[1].lower() == ".xlsx"):
            state["uploaded_template"] = stored_name
            state["uploaded_original"] = str(
                saved.get("uploaded_original") or stored_name)

        last_path = str(saved.get("last_generated_excel") or "").strip()
        if last_path:
            if not os.path.isabs(last_path):
                last_path = os.path.abspath(os.path.join(BASE, last_path))
            if os.path.splitext(last_path)[1].lower() == ".xlsx":
                state["last_generated_excel"] = last_path

        if isinstance(saved.get("continue_previous"), bool):
            state["continue_previous"] = saved["continue_previous"]
    except FileNotFoundError:
        pass
    except Exception as exc:
        log_error("读取 Excel 模板状态失败", exc)
    return state


def _save_excel_state(state):
    """原子保存模板状态；保存失败不影响当次文件生成。"""
    temp_path = EXCEL_STATE_PATH + ".tmp"
    payload = {
        "uploaded_template": state.get("uploaded_template", ""),
        "uploaded_original": state.get("uploaded_original", ""),
        "last_generated_excel": state.get("last_generated_excel", ""),
        "continue_previous": bool(state.get("continue_previous", True)),
    }
    try:
        with open(temp_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        os.replace(temp_path, EXCEL_STATE_PATH)
    except Exception as exc:
        log_error("保存 Excel 模板状态失败", exc)


def _default_excel_template(cfg):
    configured = cfg.get("excel", {}).get("template", "")
    return os.path.abspath(os.path.join(BASE, configured))


def _uploaded_excel_template(state):
    name = state.get("uploaded_template", "")
    if not name or name != os.path.basename(name):
        return ""
    return os.path.abspath(os.path.join(EXCEL_TEMPLATE_DIR, name))


def _usable_excel_template(path):
    return bool(path and os.path.isfile(path) and
                os.path.splitext(path)[1].lower() == ".xlsx")


def _select_excel_template(cfg, state, continue_previous):
    """选择本次基础表：上次结果 > 用户上传模板 > 内置模板。"""
    last_path = state.get("last_generated_excel", "")
    if continue_previous and _usable_excel_template(last_path):
        return last_path, "last_generated"

    uploaded_path = _uploaded_excel_template(state)
    if _usable_excel_template(uploaded_path):
        return uploaded_path, "uploaded"
    return _default_excel_template(cfg), "default"


def _excel_template_status(cfg, state):
    """返回前端显示所需信息，不暴露本机完整路径。"""
    default_path = _default_excel_template(cfg)
    uploaded_path = _uploaded_excel_template(state)
    has_uploaded = _usable_excel_template(uploaded_path)
    base_name = (state.get("uploaded_original") or
                 os.path.basename(uploaded_path)) if has_uploaded else os.path.basename(default_path)
    base_source = "用户上传模板" if has_uploaded else "程序内置模板"

    last_path = state.get("last_generated_excel", "")
    last_name = os.path.basename(last_path) if _usable_excel_template(last_path) else ""
    return {
        "default_name": os.path.basename(default_path),
        "base_name": base_name,
        "base_source": "uploaded" if has_uploaded else "default",
        "base_source_label": base_source,
        "last_generated_name": last_name,
        "continue_previous": bool(state.get("continue_previous", True)),
    }


def _validate_excel_template(path):
    """确认上传文件是 openpyxl 可读取且至少含一个工作表的 xlsx。"""
    from openpyxl import load_workbook

    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        if not workbook.sheetnames:
            raise ValueError("工作簿中没有工作表")
    finally:
        if workbook is not None:
            workbook.close()


def _compose_word_engineering_trend(text_map):
    """把三个现场选项追加到 Word 的工程动向正文中。"""
    details = []
    for key, label in (
            ("safety_notice", "是否安全告知"),
            ("site_plan_provided", "施工方是否提供平面图"),
            ("cooperation_level", "预估施工方配合程度")):
        value = str(text_map.get(key) or "").strip()
        if value:
            details.append("%s：%s" % (label, value))

    body = str(text_map.get("engineering_trend") or "").strip()
    if not details:
        return body
    detail_line = "；".join(details) + "。"
    return body + ("\n" if body else "") + detail_line


def _date_folder_component(value):
    text = str(value or "").strip()
    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", text)
    if match:
        return "%s-%02d-%02d" % (
            match.group(1), int(match.group(2)), int(match.group(3)))
    return datetime.datetime.now().strftime("%Y-%m-%d")


def create_output_folder(requested_root, fields):
    """在用户选择的目录下新建「日期-线路名称-设备主人」文件夹。"""
    root = str(requested_root or "").strip() or OUTPUT_DIR
    root = os.path.abspath(root)
    os.makedirs(root, exist_ok=True)
    folder_name = "-".join([
        _date_folder_component(fields.get("form_date")),
        _safe_folder_component(fields.get("line_name"), "未填线路"),
        _safe_folder_component(fields.get("equipment_owner"), "未填设备主人"),
    ])
    candidate = os.path.join(root, folder_name)
    suffix = 2
    while os.path.exists(candidate):
        candidate = os.path.join(root, "%s-%d" % (folder_name, suffix))
        suffix += 1
    os.makedirs(candidate)
    return candidate


def create_app():
    app = Flask(__name__, static_folder=STATIC_DIR, static_url_path="/static")
    app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 单次上传上限 200MB
    cfg = load_config()
    excel_state = _load_excel_state()

    # ---------- 页面 ----------
    @app.route("/")
    def index():
        return send_from_directory(STATIC_DIR, "index.html")

    # ---------- 配置 ----------
    @app.route("/api/config")
    def api_config():
        with EXCEL_STATE_LOCK:
            excel_template = _excel_template_status(cfg, excel_state)
        return jsonify({
            "title": cfg["app"].get("title", "一键流转单生成工具"),
            "fields": cfg.get("fields", []),
            "images": cfg.get("images", []),
            "attachments": cfg.get("attachments", []),
            "word_enabled": cfg.get("word", {}).get("enabled", True),
            "excel_enabled": cfg.get("excel", {}).get("enabled", False),
            "excel_template": excel_template,
            "ocr_available": ocr_engine.is_available(),
        })

    # ---------- 输出目录 ----------
    @app.route("/api/select_output_dir")
    def api_select_output_dir():
        try:
            return jsonify({"ok": True, "path": choose_output_directory()})
        except Exception as e:
            log_error("选择输出目录失败", e)
            return jsonify({"ok": False, "error": "无法打开目录选择框：%s" % e}), 500

    # ---------- OCR 识别 ----------
    @app.route("/api/ocr", methods=["POST"])
    def api_ocr():
        if "image" not in request.files:
            return jsonify({"ok": False, "error": "未收到图片文件"}), 400

        file = request.files["image"]
        ext = os.path.splitext(file.filename)[1].lower() or ".png"
        name = "img_" + uuid.uuid4().hex[:10] + ext
        path = os.path.join(UPLOAD_DIR, name)
        file.save(path)

        try:
            items = ocr_engine.recognize(path)
            extracted = ocr_engine.auto_extract(items, cfg.get("fields", []))
        except Exception as e:
            log_error("OCR 识别异常", e)
            return jsonify({"ok": False, "error": "OCR 识别失败：%s" % e,
                            "image": name}), 500

        return jsonify({
            "ok": True,
            "image": name,
            "texts": [it["text"] for it in items],
            "items": items,
            "extracted": extracted,
        })

    # ---------- 附件上传 ----------
    @app.route("/api/upload_attachment", methods=["POST"])
    def api_upload_attachment():
        if "file" not in request.files:
            return jsonify({"ok": False, "error": "未收到文件"}), 400
        file = request.files["file"]
        ext = os.path.splitext(file.filename)[1]
        name = "att_" + uuid.uuid4().hex[:10] + ext
        path = os.path.join(ATTACH_DIR, name)
        file.save(path)
        return jsonify({"ok": True, "filename": name,
                        "original": file.filename})

    # ---------- Excel 模板上传与迭代状态 ----------
    @app.route("/api/upload_excel_template", methods=["POST"])
    def api_upload_excel_template():
        if "file" not in request.files:
            return jsonify({"ok": False, "error": "未收到 Excel 模板"}), 400
        file = request.files["file"]
        original = str(file.filename or "").replace("\\", "/").rsplit("/", 1)[-1]
        if os.path.splitext(original)[1].lower() != ".xlsx":
            return jsonify({
                "ok": False,
                "error": "仅支持 .xlsx 格式，请先在 Excel 中另存为 xlsx",
            }), 400

        stem = os.path.splitext(original)[0]
        safe_stem = _safe_folder_component(stem, "Excel模板")[:60]
        stored_name = "excel_%s_%s.xlsx" % (uuid.uuid4().hex[:10], safe_stem)
        path = os.path.join(EXCEL_TEMPLATE_DIR, stored_name)
        try:
            file.save(path)
            _validate_excel_template(path)
        except Exception as exc:
            try:
                if os.path.isfile(path):
                    os.remove(path)
            except OSError:
                pass
            log_error("检查用户上传的 Excel 模板失败", exc)
            return jsonify({
                "ok": False,
                "error": "该文件不是可读取的 xlsx 工作簿：%s" % exc,
            }), 400

        with EXCEL_STATE_LOCK:
            excel_state["uploaded_template"] = stored_name
            excel_state["uploaded_original"] = original
            excel_state["last_generated_excel"] = ""
            posted_continue = request.form.get("continue_previous")
            if posted_continue in ("true", "false"):
                excel_state["continue_previous"] = posted_continue == "true"
            _save_excel_state(excel_state)
            status = _excel_template_status(cfg, excel_state)
        return jsonify({
            "ok": True,
            "original": original,
            "excel_template": status,
        })

    @app.route("/api/reset_excel_template", methods=["POST"])
    def api_reset_excel_template():
        with EXCEL_STATE_LOCK:
            keep_continue = bool(excel_state.get("continue_previous", True))
            excel_state.clear()
            excel_state.update(_empty_excel_state())
            excel_state["continue_previous"] = keep_continue
            _save_excel_state(excel_state)
            status = _excel_template_status(cfg, excel_state)
        return jsonify({"ok": True, "excel_template": status})

    # ---------- 生成文件 ----------
    @app.route("/api/generate", methods=["POST"])
    def api_generate():
        global LAST_OUTPUT_DIR
        # Word/Excel 依赖按需加载；打开程序和进行 OCR 时不必先加载 python-docx。
        from template_filler import TemplateFiller

        data = request.get_json(force=True)
        fields = data.get("fields", {})
        image_name = data.get("image_name")  # 铭牌图片文件名（uploads 下）
        attachments = data.get("attachments", [])  # [{"name","original"}, ...]

        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        outputs = []
        att_notes = []

        try:
            output_folder = create_output_folder(data.get("output_dir"), fields)
        except Exception as e:
            log_error("创建输出文件夹失败", e)
            return jsonify({"ok": False, "error": "无法创建输出文件夹：%s" % e}), 400
        LAST_OUTPUT_DIR = output_folder

        # 文本映射：所有字段 -> {{key}}
        text_map = {}
        for f in cfg.get("fields", []):
            key = f.get("key")
            text_map[key] = fields.get(key, "")
        if not text_map.get("responsible_info"):
            # Word 的 b.负责人信息优先使用项目经理2（即施工铭牌上的负责人）。
            name = text_map.get("manager2", "") or text_map.get("manager", "")
            phone = (text_map.get("manager2_phone", "") or
                     text_map.get("manager_phone", ""))
            text_map["responsible_info"] = (str(name) + " " + str(phone)).strip()
        word_text_map = dict(text_map)
        word_text_map["engineering_trend"] = _compose_word_engineering_trend(text_map)

        # 图片映射：即使未上传铭牌照片，也传入占位符列表，便于清理模板占位符。
        image_map = {
            img_cfg["placeholder"]: []
            for img_cfg in cfg.get("images", [])
            if img_cfg.get("placeholder")
        }
        if image_name:
            image_path = os.path.join(UPLOAD_DIR, image_name)
            if os.path.isfile(image_path):
                for img_cfg in cfg.get("images", []):
                    placeholder = img_cfg.get("placeholder")
                    if placeholder:
                        image_map[placeholder] = [
                            (image_path, img_cfg.get("width_cm", 8.0))]

        # 附件分类：图片类直接插入 Word，其它类型记录文件名
        att_imgs = []
        for a in attachments:
            if isinstance(a, dict):
                fn, orig = a.get("name", ""), a.get("original", "")
            else:
                fn, orig = a, a
            p = os.path.join(ATTACH_DIR, fn)
            if os.path.isfile(p):
                ext = os.path.splitext(fn)[1].lower()
                if ext in IMAGE_EXTS:
                    att_imgs.append((p, 16.0))
                else:
                    att_notes.append(orig or fn)
        attachment_text = "、".join(att_notes) if att_notes else ""

        # 生成 Word
        if cfg.get("word", {}).get("enabled"):
            w_cfg = cfg["word"]
            template = os.path.join(BASE, w_cfg["template"])
            if not os.path.isfile(template):
                return jsonify({"ok": False,
                                "error": "Word 模板不存在：" + template}), 400
            out_name = "%s_%s.docx" % (w_cfg.get("output_prefix", "文档"), stamp)
            out_path = os.path.join(output_folder, out_name)
            try:
                TemplateFiller.fill_word(
                    template, out_path, word_text_map,
                    image_map, att_imgs, attachment_text,
                    w_cfg.get("word_bindings",
                              w_cfg.get("highlight_bindings", {})))
            except Exception as exc:
                log_error("生成 Word 失败", exc)
                return jsonify({
                    "ok": False,
                    "error": "Word 生成失败，详细信息已写入崩溃日志：%s" % exc,
                }), 500
            token = uuid.uuid4().hex
            OUTPUT_FILES[token] = out_path
            outputs.append({"type": "word", "name": out_name,
                            "url": "/api/download/" + token})

        # 生成 Excel
        excel_template_info = None
        if cfg.get("excel", {}).get("enabled"):
            e_cfg = cfg["excel"]
            out_name = "%s_%s.xlsx" % (e_cfg.get("output_prefix", "表格"), stamp)
            out_path = os.path.join(output_folder, out_name)
            sheet = e_cfg.get("sheet_name") or None
            continue_excel = bool(data.get("continue_excel", True))
            try:
                with EXCEL_STATE_LOCK:
                    excel_state["continue_previous"] = continue_excel
                    template, _ = _select_excel_template(
                        cfg, excel_state, continue_excel)
                    if not _usable_excel_template(template):
                        return jsonify({"ok": False,
                                        "error": "Excel 模板不存在或已被移动"}), 400
                    if e_cfg.get("append_row"):
                        TemplateFiller.append_excel_row(
                            template, out_path, text_map, sheet,
                            e_cfg.get("column_map", {}),
                            owner_name=text_map.get("equipment_owner", ""),
                            total_sheet_name=e_cfg.get(
                                "total_sheet", "危险源统计表（总表）"))
                    else:
                        TemplateFiller.fill_excel(
                            template, out_path, text_map, sheet)
                    excel_state["last_generated_excel"] = os.path.abspath(out_path)
                    _save_excel_state(excel_state)
                    excel_template_info = _excel_template_status(cfg, excel_state)
            except Exception as exc:
                log_error("生成 Excel 失败", exc)
                return jsonify({
                    "ok": False,
                    "error": "Excel 生成失败，详细信息已写入崩溃日志：%s" % exc,
                }), 500
            token = uuid.uuid4().hex
            OUTPUT_FILES[token] = out_path
            outputs.append({"type": "excel", "name": out_name,
                            "url": "/api/download/" + token})

        return jsonify({"ok": True, "outputs": outputs,
                        "att_notes": att_notes,
                        "output_folder": output_folder,
                        "excel_template": excel_template_info})

    # ---------- 文件服务 ----------
    @app.route("/uploads/<path:filename>")
    def uploads(filename):
        return send_from_directory(UPLOAD_DIR, filename)

    @app.route("/uploads/attachments/<path:filename>")
    def attachments_file(filename):
        return send_from_directory(ATTACH_DIR, filename)

    @app.route("/output/<path:filename>")
    def download_output(filename):
        return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)

    @app.route("/api/download/<token>")
    def download_generated_file(token):
        path = OUTPUT_FILES.get(token)
        if not path or not os.path.isfile(path):
            return jsonify({"ok": False, "error": "输出文件不存在或已失效"}), 404
        return send_file(path, as_attachment=True,
                         download_name=os.path.basename(path))

    # ---------- 打开输出目录 ----------
    @app.route("/api/open_output_folder")
    def open_output_folder():
        try:
            target = LAST_OUTPUT_DIR if os.path.isdir(LAST_OUTPUT_DIR) else OUTPUT_DIR
            if os.name == "nt":
                os.startfile(target)  # noqa
            else:
                import subprocess
                subprocess.Popen(["xdg-open", target])
        except Exception:
            pass
        return jsonify({"ok": True})

    return app


def run_server(port):
    """在独立线程中启动 Flask（供 main.py 调用）。"""
    app = create_app()
    app.run(host="127.0.0.1", port=port, debug=False,
            use_reloader=False, threaded=True)


if __name__ == "__main__":
    # 开发调试入口：python app.py --web 用浏览器打开
    import threading
    import time
    import socket
    import webbrowser

    cfg = load_config()
    port = cfg["app"].get("port", 8090)
    url = "http://127.0.0.1:%d/" % port

    server = threading.Thread(target=run_server, args=(port,), daemon=True)
    server.start()
    for _ in range(50):
        try:
            with socket.create_connection(("127.0.0.1", port), timeout=0.2):
                break
        except OSError:
            time.sleep(0.1)

    webbrowser.open(url)
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        pass
